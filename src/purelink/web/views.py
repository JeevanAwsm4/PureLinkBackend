import json
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from web.models import *
from web.forms import DonorForm,Wantbloodform
from django.shortcuts import render 
from rest_framework.views import APIView 
from . models import *
from rest_framework.response import Response 
from . serializer import *
from random import randint
from openpyxl import load_workbook
import os

def get_excel_path():
    static_dir = os.path.join(BASE_DIR, 'static')
    excel_file_path = os.path.join(static_dir, 'data', 'Hospitals.xlsx')
    return excel_file_path

def index(request):
    form = DonorForm()
    return render(request,'index.html', {'form': form})


def empty(request):
    return render(request,'empty.html')

def donate_blood(request):
    if request.method == 'POST':
        form = DonorForm(request.POST)
        if form.is_valid():
            phone_number = form.cleaned_data['phone_no']
            
            # Check if a user with the given phone number already exists
            existing_user = Donor.objects.filter(phone_no=phone_number).first()
            
            if existing_user:
                # User already registered, provide a message
                response_data = {
                    "status": "error",
                    "value": "error",
                    'title': "Already Registered",
                    "message": "A user with this phone number is already registered.",
                    'redirect':'already_registered/'
                }
            else:
                # User not registered, save the form
                form.save()
                
                response_data = {
                    "status": "info",
                    "value": "success",
                    'title': "An OTP has been sent",
                    "message": "Check your messages",
                }
            
            return HttpResponse(json.dumps(response_data), content_type="application/json")
    
    else:
        form = DonorForm()

    return render(request, 'index.html', {'form': form})


def registered(request):
    return render(request, 'already.html')


def check_phone_number(request):
    if request.method == 'POST':
        phone_number = request.POST.get('phoneNumber')

        # Check if a donor with the given phone number exists
        donor = get_object_or_404(Donor, phone_no=phone_number)

        # If a donor is found, return success response

        otp = '999'
        request.session['otp'] = otp

        response_data = {
            "status": "info",
            "value": "success",
            'title': "An OTP has been sent",
            "message": "Check your messages!",
            'otp':otp
        }

        

        return HttpResponse(json.dumps(response_data), content_type="application/json")

    # If the request method is not POST, return an error response
    response_data = {
        "status": "error",
        "value": "error",
        'title': "Invalid Request",
        "message": "Invalid request method"
    }

    return HttpResponse(json.dumps(response_data), content_type="application/json")


def checkotp(request):
    if request.method == 'POST':
        user_entered_otp = request.POST.get('otp')

        # Retrieve stored OTP from the session
        stored_otp = request.session.get('otp', '')


        if user_entered_otp == stored_otp:
            # OTP is correct
            response_data = {
                "status": "success",
                "title":"Successfull",
                "message": "OTP verification successful",
                'redirect':'/empty/'
            }
        else:
            # OTP is incorrect
            response_data = {
                "status": "error",
                'title':'OTP failed',
                "message": "Incorrect OTP",
            }

        # Clear the stored OTP from the session
        request.session.pop('otp', None)

        return HttpResponse(json.dumps(response_data), content_type="application/json")

    # If the request method is not POST, return an error response
    response_data = {
        "status": "error",
        "message": "Invalid request method"
    }

    return HttpResponse(json.dumps(response_data), content_type="application/json")

def want_blood(request):
    if request.method == 'POST':
        hospitals = get_hospitals_from_excel()
        form = Wantbloodform(request.POST)
        if form.is_valid():
            phone_number = form.cleaned_data['phone_no']
            selected_hospital_name = request.POST.get('hospitals')
            selected_hospital = next((h for h in hospitals if h['name'] == selected_hospital_name), None)
            if selected_hospital:
                form.instance.latitude = selected_hospital['latitude']
                form.instance.longitude = selected_hospital['longitude']
                form.save()
            
            # Check if a user with the given phone number already exists
            existing_user = Donor.objects.filter(phone_no=phone_number).first()
            
            if existing_user:
                # User already registered, provide a message
                response_data = {
                    "status": "error",
                    "value": "error",
                    'title': "Already Registered",
                    "message": "A user with this phone number is already registered.",
                    'redirect':'already_registered/'
                }
            else:
                # User not registered, save the form
                form.save()
                
                response_data = {
                    "status": "info",
                    "value": "success",
                    'title': "An OTP has been sent",
                    "message": "Check your messages",
                }
            
            return HttpResponse(json.dumps(response_data), content_type="application/json")
    
    else:
        form = Wantbloodform()

    return render(request, 'want_blood.html', {'form': form, 'hospitals': hospitals})

class Toreact(APIView): 
    
    serializer_class = DonorData 
  
    def get(self, request): 
        detail = [ {"name": detail.name,"age": detail.age,"phone_no":detail.phone_no,"blood_group": detail.blood_group,"latitude":detail.latitude,"longitude":detail.longitude,"address":detail.address}  
        for detail in Donor.objects.all()] 
        return Response(detail) 
  
    def post(self, request): 
  
        serializer = DonorData(data=request.data) 
        if serializer.is_valid(raise_exception=True): 
            serializer.save() 
            return  Response(serializer.data)
def get_hospitals_from_excel():
    excel_sheet_path = get_excel_path()
    workbook = load_workbook(excel_sheet_path)
    sheet = workbook.active

    hospitals = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        hospital_name = row[0]
        latitude = row[1]
        longitude = row[2]

        hospitals.append({
            'name': hospital_name,
            'latitude': latitude,
            'longitude': longitude,
        })

    return hospitals
